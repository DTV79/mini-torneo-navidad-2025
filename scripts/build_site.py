import json
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl

EXCEL_PATH = Path("data/Mini Torneo Navidad 2025.xlsm")  # si usas .xlsm, cambia aquí
SHEET = "Liguilla Navidad 2025"

OUT_DIR = Path("site")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# --- Lectura específica de tu plantilla (según tu Excel actual) ---
# Grupos:
#   Grupo Z: E9:E11
#   Grupo Y: Q9:Q11
# Liguilla (resumen por pista):
#   Pista 1: X..AC (Ganador/Perdedor, sets ganados/perdidos, juegos ganados/perdidos)
#   Pista 2: AE..AJ
# Filas de liguilla: 23..54 (antes de la zona "Cruces directos")
LIGUILLA_ROW_START = 23
LIGUILLA_ROW_END = 54

def load_workbook_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET]

    grupo_z = [ws["E9"].value, ws["E10"].value, ws["E11"].value]
    grupo_y = [ws["Q9"].value, ws["Q10"].value, ws["Q11"].value]

    grupo_z = [t for t in grupo_z if isinstance(t, str) and t.strip()]
    grupo_y = [t for t in grupo_y if isinstance(t, str) and t.strip()]

    def parse_match_row(r, winner_col, loser_col, sets_w_col, sets_l_col, games_w_col, games_l_col):
        w = ws[f"{winner_col}{r}"].value
        l = ws[f"{loser_col}{r}"].value
        if not (isinstance(w, str) and isinstance(l, str)):
            return None

        sw = ws[f"{sets_w_col}{r}"].value
        sl = ws[f"{sets_l_col}{r}"].value
        gw = ws[f"{games_w_col}{r}"].value
        gl = ws[f"{games_l_col}{r}"].value

        if any(v is None for v in [sw, sl, gw, gl]):
            return None

        return {
            "winner": w.strip(),
            "loser": l.strip(),
            "sets_w": int(sw),
            "sets_l": int(sl),
            "games_w": int(gw),
            "games_l": int(gl),
        }

    matches = []
    for r in range(LIGUILLA_ROW_START, LIGUILLA_ROW_END + 1):
        # Pista 1
        m1 = parse_match_row(r, "X", "Y", "Z", "AA", "AB", "AC")
        if m1:
            matches.append({**m1, "stage": "Liguilla", "pista": 1})

        # Pista 2
        m2 = parse_match_row(r, "AE", "AF", "AG", "AH", "AI", "AJ")
        if m2:
            matches.append({**m2, "stage": "Liguilla", "pista": 2})

    return grupo_y, grupo_z, matches


def compute_standings(team_list, matches):
    stats = {t: {"PJ": 0, "V": 0, "D": 0, "Sets_F": 0, "Sets_C": 0, "Juegos_F": 0, "Juegos_C": 0} for t in team_list}

    for m in matches:
        w, l = m["winner"], m["loser"]
        if w not in stats or l not in stats:
            continue

        sw, sl = m["sets_w"], m["sets_l"]
        gw, gl = m["games_w"], m["games_l"]

        stats[w]["PJ"] += 1
        stats[l]["PJ"] += 1
        stats[w]["V"] += 1
        stats[l]["D"] += 1

        stats[w]["Sets_F"] += sw
        stats[w]["Sets_C"] += sl
        stats[l]["Sets_F"] += sl
        stats[l]["Sets_C"] += sw

        stats[w]["Juegos_F"] += gw
        stats[w]["Juegos_C"] += gl
        stats[l]["Juegos_F"] += gl
        stats[l]["Juegos_C"] += gw

    rows = []
    for t, s in stats.items():
        rows.append({
            "Equipo": t,
            **s,
            "Dif_Sets": s["Sets_F"] - s["Sets_C"],
            "Dif_Juegos": s["Juegos_F"] - s["Juegos_C"],
        })

    df = pd.DataFrame(rows)
    # Desempate: V, Dif_Sets, Dif_Juegos
    df = df.sort_values(
        by=["V", "Dif_Sets", "Dif_Juegos", "Sets_F", "Juegos_F", "Equipo"],
        ascending=[False, False, False, False, False, True],
        kind="mergesort",
    ).reset_index(drop=True)
    df.insert(0, "Pos", range(1, len(df) + 1))
    return df


def html_table_from_df(df, columns):
    # df: pandas dataframe
    header = "".join(f"<th>{c}</th>" for c in columns)
    rows = []
    for _, r in df.iterrows():
        rows.append("<tr>" + "".join(f"<td>{r[c]}</td>" for c in columns) + "</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody></table>"


def main():
    grupo_y, grupo_z, liguilla_matches = load_workbook_data()

    # Filtra matches por grupo comprobando si los equipos pertenecen al grupo
    set_y, set_z = set(grupo_y), set(grupo_z)
    matches_y = [m for m in liguilla_matches if m["winner"] in set_y and m["loser"] in set_y]
    matches_z = [m for m in liguilla_matches if m["winner"] in set_z and m["loser"] in set_z]

    standings_y = compute_standings(grupo_y, matches_y)
    standings_z = compute_standings(grupo_z, matches_z)

    # Cruces (aunque no estén jugados todavía)
    def pick(df, pos):
        r = df[df["Pos"] == pos]
        return None if r.empty else r.iloc[0]["Equipo"]

    z1, z2 = pick(standings_z, 1), pick(standings_z, 2)
    y1, y2 = pick(standings_y, 1), pick(standings_y, 2)

    sf1 = {"label": "SF1 (1º Z vs 2º Y)", "a": z1, "b": y2}
    sf2 = {"label": "SF2 (1º Y vs 2º Z)", "a": y1, "b": z2}

    # Ganador final: solo lo sabremos si hay final rellenada en tu Excel.
    # Como tu plantilla todavía no tiene la final “volcada” en una zona resumen,
    # dejamos campeón en "Pendiente" (y se actualizará cuando lo tengamos calculable).
    champion = "Pendiente"

    payload = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "excel_file": str(EXCEL_PATH),
        "groups": {"Y": grupo_y, "Z": grupo_z},
        "matches": liguilla_matches,
        "standings": {
            "Y": standings_y.to_dict(orient="records"),
            "Z": standings_z.to_dict(orient="records"),
        },
        "crosses": {"semifinals": [sf1, sf2], "final": {"champion": champion}},
    }

    (OUT_DIR / "standings.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Web
    def match_row(m):
        return f"<tr><td>{m['stage']}</td><td>{m['pista']}</td><td>{m['winner']}</td><td>{m['loser']}</td><td>{m['sets_w']}-{m['sets_l']}</td><td>{m['games_w']}-{m['games_l']}</td></tr>"

    matches_html = ""
    if liguilla_matches:
        matches_html = f"""
        <table>
          <thead>
            <tr><th>Fase</th><th>Pista</th><th>Ganador</th><th>Perdedor</th><th>Sets</th><th>Juegos</th></tr>
          </thead>
          <tbody>
            {''.join(match_row(m) for m in liguilla_matches)}
          </tbody>
        </table>
        """
    else:
        matches_html = "<p class='muted'>Aún no hay partidos registrados.</p>"

    standings_cols = ["Pos", "Equipo", "PJ", "V", "D", "Dif_Sets", "Dif_Juegos"]

    html = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Mini Torneo Navidad 2025</title>
  <style>
    body{{font-family:system-ui,Segoe UI,Roboto,Arial;margin:24px;max-width:1050px}}
    .grid{{display:grid;grid-template-columns:1fr;gap:14px}}
    @media (min-width: 900px){{ .grid2{{grid-template-columns:1fr 1fr}} }}
    .card{{border:1px solid #ddd;border-radius:14px;padding:16px}}
    h1{{margin:0 0 6px 0}}
    h2{{margin:0 0 10px 0;font-size:18px}}
    .muted{{color:#666;font-size:14px;margin:6px 0 0}}
    table{{border-collapse:collapse;width:100%}}
    th,td{{border-bottom:1px solid #eee;padding:8px;text-align:left;font-size:14px}}
    th{{background:#fafafa}}
    .pill{{display:inline-block;padding:6px 10px;border:1px solid #ddd;border-radius:999px;margin-right:6px}}
  </style>
</head>
<body>
  <h1>Mini Torneo Navidad 2025 🎄🎾</h1>
  <p class="muted">Última generación: {payload["updated_at"]} (se actualiza al subir la Excel al repo)</p>

  <div class="grid grid2" style="margin-top:14px">
    <div class="card">
      <h2>Grupos</h2>
      <div><span class="pill"><b>Grupo Z</b></span> {" · ".join(grupo_z) if grupo_z else "—"}</div>
      <div style="margin-top:8px"><span class="pill"><b>Grupo Y</b></span> {" · ".join(grupo_y) if grupo_y else "—"}</div>
    </div>

    <div class="card">
      <h2>Cruces</h2>
      <div><b>{sf1["label"]}:</b> {sf1["a"] or "—"} vs {sf1["b"] or "—"}</div>
      <div style="margin-top:6px"><b>{sf2["label"]}:</b> {sf2["a"] or "—"} vs {sf2["b"] or "—"}</div>
      <div style="margin-top:10px"><b>Campeón:</b> {champion}</div>
      <p class="muted">El campeón aparecerá cuando haya final calculable en la plantilla.</p>
    </div>
  </div>

  <div class="grid" style="margin-top:14px">
    <div class="card">
      <h2>Partidos jugados</h2>
      {matches_html}
    </div>

    <div class="grid grid2">
      <div class="card">
        <h2>Clasificación Grupo Z</h2>
        {html_table_from_df(standings_z, standings_cols) if not standings_z.empty else "<p class='muted'>—</p>"}
      </div>
      <div class="card">
        <h2>Clasificación Grupo Y</h2>
        {html_table_from_df(standings_y, standings_cols) if not standings_y.empty else "<p class='muted'>—</p>"}
      </div>
    </div>
  </div>

  <p class="muted" style="margin-top:14px">
    Desempates: diferencia de sets, luego diferencia de juegos.
  </p>
</body>
</html>
"""
    (OUT_DIR / "index.html").write_text(html, encoding="utf-8")


if __name__ == "__main__":
    main()
